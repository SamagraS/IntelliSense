"""
company_batch_crawler.py
========================
Batch runner that ensures all companies in the portfolio are crawled,
with at least one company per sector covered.

Builds on news_intelligence.py (NewsCrawler).

Usage:
    python company_batch_crawler.py
    python company_batch_crawler.py --mode live --save
    python company_batch_crawler.py --companies-file companies.csv --mode background --save
    python company_batch_crawler.py --list
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from news_intelligence import NewsCrawler


DEFAULT_COMPANY_REGISTRY = [
    {
        "company_id": "20microns",
        "company_name": "20 Microns Limited",
        "promoter_name": "Chandresh Parikh",
        "sector": "Manufacturing",
    },
    {
        "company_id": "360one",
        "company_name": "360 ONE WAM Limited",
        "promoter_name": "Karan Bhagat",
        "sector": "NBFC",
    },
    {
        "company_id": "5paisa",
        "company_name": "5Paisa Capital Limited",
        "promoter_name": "Prakarsh Gagdani",
        "sector": "NBFC",
    },
    {
        "company_id": "aadhar_housing",
        "company_name": "Aadhar Housing Finance Limited",
        "promoter_name": "Rishi Anand",
        "sector": "NBFC",
    },
    {
        "company_id": "aavas",
        "company_name": "Aavas Financiers Limited",
        "promoter_name": "Sushil Kumar Agarwal",
        "sector": "NBFC",
    },
    {
        "company_id": "aditya_birla_capital",
        "company_name": "Aditya Birla Capital Limited",
        "promoter_name": "Vishakha Mulye",
        "sector": "NBFC",
    },
    {
        "company_id": "aarey_drugs",
        "company_name": "Aarey Drugs & Pharmaceuticals Limited",
        "promoter_name": "Hasmukh Shah",
        "sector": "Pharma",
    },
    {
        "company_id": "aarti_drugs",
        "company_name": "Aarti Drugs Limited",
        "promoter_name": "Adhish Patil",
        "sector": "Pharma",
    },
    {
        "company_id": "aditya_birla_fashion",
        "company_name": "Aditya Birla Fashion and Retail Limited",
        "promoter_name": "Ashish Dikshit",
        "sector": "Retail",
    },
    {
        "company_id": "accelya",
        "company_name": "Accelya Solutions India Limited",
        "promoter_name": "Anand Venkataraman",
        "sector": "IT",
    },
]

COMPANY_REGISTRY = DEFAULT_COMPANY_REGISTRY


def connect_from_env():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        return None

    try:
        import psycopg2

        conn = psycopg2.connect(database_url)
        print("DB storage enabled via DATABASE_URL")
        return conn
    except Exception as exc:
        print(f"Warning: could not connect to DATABASE_URL: {exc}")
        return None


def slugify_company_id(company_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", company_name.lower()).strip("_")
    return slug or f"company_{int(time.time())}"


def normalize_sector(raw_sector: str) -> str:
    if raw_sector is None:
        return "Unknown"
    cleaned = str(raw_sector).strip()
    cleaned = re.sub(r"^\d+\s*", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned or "Unknown"


def normalize_row(row: dict, index: int) -> Optional[dict]:
    lowered = {str(key).strip().lower(): value for key, value in row.items()}

    company_name = (
        lowered.get("company_name")
        or lowered.get("name of company")
        or lowered.get("company")
        or lowered.get("name")
    )
    if not company_name:
        return None

    promoter_name = (
        lowered.get("promoter_name")
        or lowered.get("promoter")
        or lowered.get("promoter name")
        or ""
    )
    sector = (
        lowered.get("sector")
        or lowered.get("project_sector")
        or lowered.get("project sector")
        or lowered.get("industry")
        or "Unknown"
    )
    company_id = (
        lowered.get("company_id")
        or lowered.get("company id")
        or slugify_company_id(str(company_name))
    )

    return {
        "company_id": str(company_id).strip() or f"company_{index}",
        "company_name": str(company_name).strip(),
        "promoter_name": str(promoter_name).strip(),
        "sector": normalize_sector(str(sector)),
    }


def load_companies_from_csv(file_path: Path) -> list[dict]:
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        rows = [normalize_row(row, index) for index, row in enumerate(reader, 1)]
    return [row for row in rows if row]


def load_companies_from_json(file_path: Path) -> list[dict]:
    with file_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)

    if isinstance(data, dict):
        items = data.get("companies", [])
    else:
        items = data

    rows = [normalize_row(item, index) for index, item in enumerate(items, 1) if isinstance(item, dict)]
    return [row for row in rows if row]


def load_companies_from_xlsx(file_path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires openpyxl. Install it or provide CSV/JSON instead.") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    normalized = []
    for index, values in enumerate(rows[1:], 1):
        row = {headers[col]: values[col] for col in range(min(len(headers), len(values)))}
        item = normalize_row(row, index)
        if item:
            normalized.append(item)
    return normalized


def load_company_registry(companies_file: Optional[str] = None) -> list[dict]:
    if not companies_file:
        return list(DEFAULT_COMPANY_REGISTRY)

    file_path = Path(companies_file).expanduser()
    if not file_path.is_absolute():
        file_path = Path.cwd() / file_path
    if not file_path.exists():
        example_path = Path(__file__).resolve().parent / "companies_template.csv"
        raise FileNotFoundError(
            f"Companies file not found: {file_path}. "
            f"Use an existing CSV/JSON/XLSX path, or start from: {example_path}"
        )

    suffix = file_path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        companies = load_companies_from_csv(file_path)
    elif suffix == ".json":
        companies = load_companies_from_json(file_path)
    elif suffix in {".xlsx", ".xlsm"}:
        companies = load_companies_from_xlsx(file_path)
    else:
        raise ValueError("Unsupported companies file. Use CSV, JSON, or XLSX.")

    if not companies:
        raise ValueError(f"No valid companies found in {file_path}")

    print(f"Loaded {len(companies)} companies from: {file_path}")
    return companies


class BatchCrawler:
    def __init__(self, db_conn=None, delay_seconds=1.5, company_registry: Optional[list[dict]] = None):
        self.crawler = NewsCrawler(db_conn=db_conn, delay_seconds=delay_seconds)
        self.db = db_conn
        self.company_registry = company_registry or list(DEFAULT_COMPANY_REGISTRY)
        self.crawled_sectors: set[str] = set()
        self.results: dict[str, list] = {}
        self.errors: dict[str, str] = {}

    def get_companies_by_sector(self) -> dict[str, list]:
        by_sector = defaultdict(list)
        for company in self.company_registry:
            by_sector[company["sector"]].append(company)
        return dict(by_sector)

    def print_coverage_map(self):
        by_sector = self.get_companies_by_sector()
        print("\n" + "=" * 65)
        print(f"{'SECTOR':<20} {'COMPANIES':>5}  NAMES")
        print("=" * 65)
        for sector, companies in sorted(by_sector.items()):
            names = ", ".join(company["company_name"] for company in companies)
            print(f"{sector:<20} {len(companies):>5}  {names[:60]}")
        print("=" * 65)
        print(f"{'TOTAL':<20} {len(self.company_registry):>5}")
        print()

    def run_all(self, mode: str = "background", sector_filter: str = None, max_per_query: int = 5, delay_between: float = 2.0) -> dict:
        companies = self.company_registry
        if sector_filter:
            companies = [company for company in companies if company["sector"].lower() == sector_filter.lower()]
            if not companies:
                print(f"No companies found for sector: {sector_filter}")
                return {}

        print("\n" + "=" * 65)
        print(f"BATCH CRAWL - {len(companies)} companies, mode={mode.upper()}")
        if sector_filter:
            print(f"Sector filter: {sector_filter}")
        print("=" * 65)

        total_articles = 0
        start_time = time.time()

        for index, company in enumerate(companies, 1):
            cid = company["company_id"]
            cname = company["company_name"]
            pname = company["promoter_name"] or cname
            sector = company["sector"]

            print(f"\n[{index}/{len(companies)}] -- {cname} ({sector}) --")
            try:
                if mode == "background":
                    articles = self._run_background_smart(company, max_per_query=max_per_query)
                else:
                    articles = self.crawler.run_live_refresh(
                        company_name=cname,
                        promoter_name=pname,
                        company_id=cid,
                        max_per_query=max_per_query,
                    )

                self.results[cid] = articles
                total_articles += len(articles)
                print(f"  Saved {len(articles)} articles for {cname}")
            except Exception as exc:
                self.errors[cid] = str(exc)
                print(f"  ERROR for {cname}: {exc}")

            if index < len(companies):
                print(f"  (waiting {delay_between}s before next company...)")
                time.sleep(delay_between)

        elapsed = time.time() - start_time
        return self._build_summary(total_articles, elapsed)

    def _run_background_smart(self, company: dict, max_per_query: int) -> list[dict]:
        from news_intelligence import build_background_queries

        cid = company["company_id"]
        cname = company["company_name"]
        pname = company["promoter_name"] or cname
        sector = company["sector"]
        all_queries = build_background_queries(cname, pname, sector)

        if sector in self.crawled_sectors:
            queries = [query for query in all_queries if query["scope"] != "sector"]
            saved = sum(1 for query in all_queries if query["scope"] == "sector")
            print(f"  (sector '{sector}' already crawled - skipping {saved} sector queries)")
        else:
            queries = all_queries
            self.crawled_sectors.add(sector)
            print(f"  (first company in sector '{sector}' - running full sector queries)")

        all_records = []
        for query in queries:
            records = self.crawler._process_query(
                query_dict=query,
                company_id=cid,
                promoter_name=pname if query["scope"] == "promoter" else None,
                sector=sector if query["scope"] == "sector" else None,
                crawl_phase="background_deep_crawl",
                max_articles=max_per_query,
            )
            all_records.extend(records)
            self.crawler._write_to_db(records)
        return all_records

    def _build_summary(self, total_articles: int, elapsed: float) -> dict:
        by_sector = defaultdict(int)
        by_company = {}

        for company in self.company_registry:
            cid = company["company_id"]
            sector = company["sector"]
            count = len(self.results.get(cid, []))
            by_sector[sector] += count
            by_company[company["company_name"]] = count

        summary = {
            "total_articles": total_articles,
            "companies_crawled": len(self.results),
            "companies_errored": len(self.errors),
            "by_sector": dict(by_sector),
            "by_company": by_company,
            "errors": self.errors,
            "elapsed_seconds": round(elapsed, 1),
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }

        print("\n" + "=" * 65)
        print("BATCH CRAWL COMPLETE")
        print("=" * 65)
        print(f"Total articles:    {total_articles}")
        print(f"Companies crawled: {len(self.results)}")
        print(f"Errors:            {len(self.errors)}")
        print(f"Time elapsed:      {elapsed:.1f}s\n")
        print(f"{'SECTOR':<20} {'ARTICLES':>10}")
        print("-" * 32)
        for sector, count in sorted(summary["by_sector"].items()):
            print(f"{sector:<20} {count:>10}")
        print("-" * 32)
        print(f"{'TOTAL':<20} {total_articles:>10}")

        if self.errors:
            print("\nErrors:")
            for cid, err in self.errors.items():
                print(f"  {cid}: {err}")

        return summary

    def save_results_json(self, filepath: str | None = None):
        if filepath is None:
            output_dir = Path(__file__).resolve().parent / "outputs"
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"crawl_results_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
            filepath = str(output_dir / filename)

        output = []
        for articles in self.results.values():
            for article in articles:
                row = dict(article)
                row["published_date"] = str(row.get("published_date", ""))
                row["crawl_timestamp"] = str(row.get("crawl_timestamp", ""))
                output.append(row)

        with open(filepath, "w", encoding="utf-8") as handle:
            json.dump(output, handle, indent=2, ensure_ascii=False)
        print(f"\nResults saved to: {filepath} ({len(output)} articles)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Batch news crawler for credit appraisal engine")
    parser.add_argument("--sector", type=str, default=None, help="Filter to one sector (e.g. NBFC)")
    parser.add_argument("--mode", type=str, default="background", help="background | live")
    parser.add_argument("--list", action="store_true", help="Print coverage map and exit")
    parser.add_argument("--save", action="store_true", help="Save results to JSON")
    parser.add_argument("--save-path", type=str, default=None, help="Optional JSON output path")
    parser.add_argument("--delay", type=float, default=1.5, help="Seconds between article fetches")
    parser.add_argument("--companies-file", type=str, default=None, help="Optional CSV, JSON, or XLSX companies file")
    args = parser.parse_args()

    company_registry = load_company_registry(args.companies_file)
    db_conn = connect_from_env()
    batch = BatchCrawler(db_conn=db_conn, delay_seconds=args.delay, company_registry=company_registry)

    try:
        if args.list:
            batch.print_coverage_map()
        else:
            batch.print_coverage_map()
            batch.run_all(mode=args.mode, sector_filter=args.sector)
            if args.save:
                batch.save_results_json(args.save_path)
    finally:
        if db_conn:
            db_conn.close()

